from flask import Flask, request, send_file, render_template_string
from werkzeug.utils import secure_filename
import pandas as pd
import os
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

SOURCES = ['Geotab', 'Netradyne', 'Safeguard']
VIOLATION_TYPES = ['Distraction', 'Following Distance', 'Sign Signal', 'Speeding', 'Seatbelt']

UPLOAD_FORM_HTML = """
<!doctype html>
<html>
<head>
    <title>Safety E-Mail (Geotab & Netradyne)</title>
    <style>
        input[type="submit"], button {
            color: #ffffff;
            background-color: #2d63c8;
            font-size: 19px;
            border: 1px solid #2d63c8;
            padding: 12px 45px;
            cursor: pointer;
        }
        input[type="submit"]:hover, button:hover {
            color: #2d63c8;
            background-color: #ffffff;
        }
    </style>
</head>
<body>
    <h3>Upload File for Processing</h3>
    <form action="/" method="post" enctype="multipart/form-data">
        <input type="file" name="file" accept=".csv"><br><br>
        <input type="submit" value="Upload"><br><br>
        
        <h4>Select Sources:</h4>
        {% for source in sources %}
            <label><input type="checkbox" name="source" value="{{ source }}" {% if source == 'Geotab' %}checked{% endif %}> {{ source }}</label><br>
        {% endfor %}

        <h4>Select Violations:</h4>
        {% for violation in violation_types %}
            <label><input type="checkbox" name="violation" value="{{ violation }}" checked> {{ violation }}</label><br>
        {% endfor %}
    </form>
</body>
</html>
"""

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Ensure the uploads folder exists
        if not os.path.exists('uploads'):
            os.makedirs('uploads')

        file = request.files['file']
        filename = secure_filename(file.filename)
        file_path = os.path.join('uploads', filename)
        file.save(file_path)

        selected_sources = request.form.getlist('source')
        selected_violations = request.form.getlist('violation')

        df = pd.read_csv(file_path)
        processed_data = process_data(df, selected_sources, selected_violations)

        output_file = filename.replace('.csv', '_processed.xlsx')
        full_output_path = os.path.join('uploads', output_file)
        save_to_excel(processed_data, full_output_path)

        return f'''
            <!doctype html>
            <html>
            <head>
                <style>
                    button {{
                        color: #ffffff;
                        background-color: #2d63c8;
                        font-size: 19px;
                        border: 1px solid #2d63c8;
                        padding: 12px 45px;
                        cursor: pointer;
                    }}
                    button:hover {{
                        color: #2d63c8;
                        background-color: #ffffff;
                    }}
                </style>
            </head>
            <body>
                <a href="/download/{output_file}">
                    <button>Download Processed File</button>
                </a>
            </body>
            </html>
        '''

    return render_template_string(UPLOAD_FORM_HTML, sources=SOURCES, violation_types=VIOLATION_TYPES)

def process_data(df, selected_sources, selected_violations):
    corrected_column_name = 'Delivery Associate '  # Adjust if the column name is different
    df = df[df['Source'].isin(selected_sources) & df['Metric Type'].isin(selected_violations)]
    grouped = df.groupby([corrected_column_name, 'Source'])
    result = []

    for (name, source), group in grouped:
        violations = group['Metric Type'].value_counts()
        violations_str = ', '.join([f"{v} ({c})" for v, c in violations.items()])
        result.append([name, violations_str, source])

    extracted_data = pd.DataFrame(result, columns=['Name', 'Violations', 'Source'])
    return extracted_data

def save_to_excel(df, filename):
    # Ensure the directory exists
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    
    # Save the DataFrame to an Excel file
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Processed Data')
        
        # Get the workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets['Processed Data']

        # Set alignment and border for all cells
        for row in worksheet.iter_rows(min_row=1, max_col=len(df.columns), max_row=len(df)+1):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                     top=Side(style='thin'), bottom=Side(style='thin'))

        # Set fill color for header cells
        fill = PatternFill(fill_type='solid', fgColor='B8CCE4')
        for cell in worksheet[1]:
            cell.fill = fill
            cell.font = Font(bold=True)

        # Adjust column widths
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells) + 2  # Added 2 for extra spacing
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

@app.route('/download/<filename>')
def download(filename):
    file_path = os.path.join('uploads', filename)
    return send_file(file_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
